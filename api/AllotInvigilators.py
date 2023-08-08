import openpyxl
from PhDStudent import PhDStudent 
from Course import Course
import pandas as pd
import math

def extract_course_code(course_code):
    for i in range(0,len(course_code)):
            if course_code[i][0].isnumeric():
                str = ''
                for j in course_code[i-1]:
                    if j.isnumeric() :
                        break
                    str += j    
                course_code[i] = str+course_code[i]
    return course_code

def hasExam(course_pool, course_code):
    course = course_pool[course_code]
    return course.get_date()

def removeFromPool(available_pool, available_sorted_keys, student):
    if student.get_admission_no() in available_pool:
        del available_pool[student.get_admission_no()]
    if student.get_admission_no() in available_sorted_keys:
        available_sorted_keys.remove(student.get_admission_no())

def sameDayDuty(student, date):
    duties = student.get_duties()
    for duty in duties:
        duty_date = duty.get_date()
        if date == duty_date:
            return True
    return False

def clearAllotedList(alloted_students_pool, students_available_copy):
    for student in students_available_copy:
        if student in alloted_students_pool:
            del alloted_students_pool[student]

def get_num_invigilator(strength):
        result = strength / 25
        if strength >= 150 or strength % 25 >= 12:
            return math.ceil(result)
        else:
            return math.floor(result)

room_capacity = {"c01" : 106, "c02" : 42, "c03" : 18, "c11" : 94, "c12" : 39, "c13" : 18, "c21" : 80, "c22" : 18, "c24" : 18,
                 "c101" : 169, "c102" : 281, "c201" : 177, "c208" : 20, "c209" : 22, "c210" : 25, "c211" : 24, "c212" : 28,
                 "c213" : 31, "c214" : 20, "c215" : 22, "c216" : 17, "a006" : 58, "a007" : 56, "a106" : 48, "b105" : 42}

#########################################################################################################################
########################## Extracting PhD Students from Student Course Registration Excel File ##########################
#########################################################################################################################

# List of Students from "StudentRegistration" Excel File.
phd_students = {}

# Relative Path of the Excel file
file_path = 'StudentRegistration.xlsx'

# Names of the relevant Columns in the Excel file
admission_no_col = 'Admission No.'
name_col = 'Name'
email_col = 'Email Id'
course_code_col = 'Course Code'

# Open the Excel file
workbook = openpyxl.load_workbook(filename=file_path)

# Slecting the First(0th - indexed) Worksheet/Sheet in the Workbook/Excel File.
worksheet = workbook.worksheets[0]


# Geting the Column Indices based on the Column Names
header_row = worksheet[1]
column_names = tuple(cell.value for cell in header_row)

try:
    admission_no_col_idx = column_names.index(admission_no_col) + 1
    name_col_idx = column_names.index(name_col) + 1
    email_col_idx = column_names.index(email_col) + 1
    course_code_col_idx = column_names.index(course_code_col) + 1
except ValueError as e:
    print("Error: One or more Required Columns not found in the Excel File!")


# Iterate over rows and extract data from relevant Columns.
for row in worksheet.iter_rows(min_row=2, values_only=True):
    admission_no = row[admission_no_col_idx - 1]
    name = row[name_col_idx - 1]
    email = row[email_col_idx - 1]
    course_code = row[course_code_col_idx - 1]


    if any(val is None for val in [admission_no, name, email, course_code]):
            break
    admission_no = admission_no.lower()
    course_code = course_code.lower()
    # print(admission_no, course_code)
    if admission_no not in phd_students:
            phd_students[admission_no] = PhDStudent(admission_no, name, email)
    # Add the enrolled course to the PhDStudent object
    phd_students[admission_no].add_enrolled_course(course_code)


#########################################################################################################################
##################################### Extracting Course List from Exam Date Sheet  ######################################
#########################################################################################################################

course_pool = {}
course_list = []

# Relative Path of the Excel file
file_path = 'ExamDateSheet.xlsx'

# Names of the relevant Columns in the Excel file
date_col = 'Date'
day_col = 'Day'
time_col = 'Time'
accronynm_col = 'Accronynm'
course_code_col = 'Course Code'
strength_col = 'Strength'
room_no_col = 'Room No.'

# Open the Excel file
workbook = openpyxl.load_workbook(filename=file_path)

# Slecting the First(0th - indexed) Worksheet/Sheet in the Workbook/Excel File.
worksheet = workbook.worksheets[0]


# Geting the Column Indices based on the Column Names
header_row = worksheet[1]
column_names = tuple(cell.value for cell in header_row)
# course_code = [str.strip() for str in course_code]


try:
    date_col_idx = column_names.index(date_col) + 1
    day_col_idx = column_names.index(day_col) + 1
    time_col_idx = column_names.index(time_col) + 1
    accronynm_col_idx = column_names.index(accronynm_col) + 1
    course_code_col_idx = column_names.index(course_code_col) + 1
    strength_col_idx = column_names.index(strength_col) + 1
    room_no_col_idx = column_names.index(room_no_col) + 1
except ValueError as e:
    print("Error: One or more Required Columns not found in the Excel File!")

# Iterate over rows and extract data from relevant Columns.
for row in worksheet.iter_rows(min_row=2, values_only=True):

    date = row[date_col_idx - 1]
    day = row[day_col_idx - 1]
    time = row[time_col_idx - 1]
    accronynm = row[accronynm_col_idx - 1]
    course_code = row[course_code_col_idx - 1]
    strength = row[strength_col_idx - 1]
    room_no = row[room_no_col_idx - 1]

    if any(val is None for val in [date, day, time, accronynm, course_code, strength, room_no]):
        break

    # Making the Course List from the DateSheet.
    # Spiltting the Course Code as there can be multiple Course Code for a Course.
    course_code = course_code.split("/")

    # Removing leading and trailing spaces from each substring
    course_code = [str.strip() for str in course_code]
    course_code = extract_course_code(course_code)
    course_code = [s.lower() for s in course_code]
    new_course = Course(accronynm, strength, date, time, day, room_no, room_capacity)
    course_list.append(new_course)
    for code in course_code:
        course_pool[code] = new_course



#########################################################################################################################
##################################### Loading Eligible Students for Invilgilation Duty ##################################
#########################################################################################################################

students_pool = {}
students_previous = {}
students_available = {}




# Relative Path of the Excel file
file_path = 'StudentList.xlsx'

# Names of the relevant Columns in the Excel file
admission_no_col = 'Admission No.'
name_col = 'Name'
email_col = 'Email ID'
available_col = 'Available'
previous_col = 'Previous'

# Open the Excel file
workbook = openpyxl.load_workbook(filename=file_path)

# Slecting the First(0th - indexed) Worksheet/Sheet in the Workbook/Excel File.
worksheet = workbook.worksheets[0]

# Geting the Column Indices based on the Column Names
header_row = worksheet[1]
column_names = tuple(cell.value for cell in header_row)


try:
    admission_no_col_idx = column_names.index(admission_no_col) + 1
    name_col_idx = column_names.index(name_col) + 1
    email_col_idx = column_names.index(email_col) + 1
    available_col_idx = column_names.index(available_col) + 1
    previous_col_idx = column_names.index(previous_col) + 1
except ValueError as e:
    print("Error: One or more Required Columns not found in the Excel File!")

# Iterate over rows and extract data from relevant Columns.
for row in worksheet.iter_rows(min_row=2, values_only=True):
    admission_no = row[admission_no_col_idx - 1]
    name = row[name_col_idx - 1]
    email = row[email_col_idx - 1]
    available = row[available_col_idx - 1]
    previous = row[previous_col_idx - 1]
    
    
    if any(val is None for val in [admission_no, name, email, available, previous]):
        break
    admission_no = admission_no.lower()

    if(admission_no not in phd_students):
        phd_students[admission_no] = PhDStudent(admission_no, name, email)
    if available.lower() == "yes":
        students_pool[admission_no] = phd_students[admission_no]
    if(available.lower() == "yes" and previous.lower() == "yes"):
        students_previous[admission_no] = phd_students[admission_no]
    elif(available.lower() == "yes"):
        students_available[admission_no] = phd_students[admission_no]

students_available_copy = students_available.copy()
students_previous_copy = students_previous.copy()


#########################################################################################################################
############################################### Extracting Course TAs ###################################################
#########################################################################################################################

# Relative Path of the Excel file
file_path = 'TAList.xlsx'

# Names of the relevant Columns in the Excel file
admission_no_col = 'Admission No.'
email_col = 'Email ID'
course_code_col = 'Course Code'
remark_col = 'Remark'
name_col = 'Name'


# Open the Excel file
workbook = openpyxl.load_workbook(filename=file_path)

# Slecting the First(0th - indexed) Worksheet/Sheet in the Workbook/Excel File.
worksheet = workbook.worksheets[0]

# Geting the Column Indices based on the Column Names
header_row = worksheet[1]
column_names = tuple(cell.value for cell in header_row)


try:
    admission_no_col_idx = column_names.index(admission_no_col) + 1
    name_col_idx = column_names.index(name_col) + 1
    email_col_idx = column_names.index(email_col) + 1
    course_code_col_idx = column_names.index(course_code_col) + 1
    remark_col_idx = column_names.index(remark_col) + 1
except ValueError as e:
    print("Error: One or more Required Columns not found in the Excel File!")

# Iterate over rows and extract data from relevant Columns.
for row in worksheet.iter_rows(min_row=2, values_only=True):
    admission_no = row[admission_no_col_idx - 1]
    name = row[name_col_idx - 1]
    email = row[email_col_idx - 1]
    course_code = row[course_code_col_idx - 1]
    remark = row[remark_col_idx - 1]

    try:
        if any(val is None for val in [admission_no, name, email]):
            break
    except:
        print("Error: Make sure the Columns(Admission No., Name, Email ID) are Filled")

    

    admission_no = admission_no.lower()
    if remark.lower() == "yes":
        course_code = course_code.split("_")[1]
        course_code = course_code.split("/")
        course_code = [str.strip() for str in course_code]
        course_code = extract_course_code(course_code)
        course_code = [s.lower() for s in course_code]
        for code in course_code:
            if code in course_pool:
                curr_course = course_pool[code]
                if(admission_no in students_pool):
                    student = students_pool[admission_no]
                    curr_course.add_course_ta(student)
                    break


#########################################################################################################################
#################################################### Allocation #########################################################
#########################################################################################################################

alloted_students_pool = {}

# Reverse Sorting the Course List based on the Strength of the Course.
course_list = sorted(course_list, key=lambda x: x.get_strength(), reverse=True)

# Sorting the Keys of Available Students in Decending Order. 
available_sorted_keys = sorted(students_available, key=lambda x: x, reverse=True)

# Sorting the Keys of Previous Students in Decending Order.
previous_sorted_keys = sorted(students_previous, key=lambda x: x, reverse=True)


# Alloting duties to all the TAs for their respective Courses
for course in course_list:
    date = course.get_date()
    
    req_invigilators = course.get_req_invigilator()
    if(req_invigilators == 0):
        continue

    course_ta = course.get_course_ta()
    for ta in course_ta:
        if req_invigilators == len(course.get_invigilators()):
            break
        if ta.get_admission_no() in students_pool:
            ta_course = ta.get_enrolled_courses()
            can_allot = True
            for ta_course_code in ta_course:
                if (ta_course_code in course_pool) and (hasExam(course_pool, ta_course_code) == date):
                    can_allot = False
                    break
            if can_allot:
                course.add_invigilators(ta)
                removeFromPool(students_available, available_sorted_keys, ta)
                alloted_students_pool[ta.get_admission_no()] = ta
                ta.add_duty(course)

for course in course_list:
    date = course.get_date()
    req_invigilators = course.get_req_invigilator()
    if(req_invigilators == 0 or len(course.get_invigilators()) == req_invigilators):
        continue
    # print(course.get_course_acronym())
    full_allotment = False
    while(full_allotment == False):
        # print(f"Available Students: {len(available_sorted_keys)}")
        # print(f"Alloted Students: {len(alloted_students_pool)}")
        for student in available_sorted_keys:
            if req_invigilators == len(course.get_invigilators()):
                full_allotment = True
                break
            student_course = students_available[student].get_enrolled_courses()
            can_allot = True
            for student_course_code in student_course:
                if(((student_course_code in course_pool) and (hasExam(course_pool, student_course_code) == date)) or (sameDayDuty(students_available[student], date))):
                    can_allot = False
                    break
            if can_allot:
                students_available[student].add_duty(course)
                course.add_invigilators(students_available[student])
                removeFromPool(students_available, available_sorted_keys, students_available[student])
                alloted_students_pool[student] = students_pool[student]

        # print(f"Available Students: {len(available_sorted_keys)}")
        # print(f"Alloted Students: {len(alloted_students_pool)}")       
        if(len(available_sorted_keys) == 0):
            clearAllotedList(alloted_students_pool, students_available_copy)
            # print("It was Here!")
            # print(f"Updated Alloted Students: {len(alloted_students_pool)}")

        if full_allotment == False:
            for student in previous_sorted_keys:
                if req_invigilators == len(course.get_invigilators()):
                    full_allotment = True
                    break
                student_course = students_previous[student].get_enrolled_courses()
                can_allot = True
                for student_course_code in student_course:
                    if(((student_course_code in course_pool) and (hasExam(course_pool, student_course_code) == date)) or (sameDayDuty(students_previous[student], date))):
                        can_allot = False
                        break
                if can_allot:
                    students_previous[student].add_duty(course)
                    course.add_invigilators(students_previous[student])
                    removeFromPool(students_previous, previous_sorted_keys, students_previous[student])
                    alloted_students_pool[student] = students_pool[student]
        # print(f"Alloted Students: {len(alloted_students_pool)}")
        if(len(available_sorted_keys) == 0 and len(previous_sorted_keys) == 0):
            students_available = students_available_copy.copy()
            available_sorted_keys = sorted(students_available, key=lambda x: x, reverse=True)
            students_previous = students_previous_copy.copy()
            previous_sorted_keys = sorted(students_previous, key=lambda x: x, reverse=True)

if(len(available_sorted_keys) == 0):
    clearAllotedList(alloted_students_pool, students_available_copy)
    


            
# # for course in course_list:
# #     date = course.get_date()
# #     print(course.get_course_acronym())
# #     print(f"Date: {date}")
# #     req_invigilators = course.get_req_invigilator()
# #     print(f"Req_Invigilators: {req_invigilators}")
# #     invigilators = course.get_invigilators()
# #     print(f"Alloted Invigilators: {len(invigilators)}")
# #     # print(len(invigilators))
# #     for student in invigilators:
# #         print(f"Admission No. {student.get_admission_no()} Name: {student.get_name()}")
# #         # for course in student.get_enrolled_courses():
# #         #     if course in course_pool:
# #         #         course = course_pool[course]
# #         #         print(f" Student Course: {course.get_course_acronym()} Date: {course.get_date()}")

# # for student in students_pool:
# #     print(f"Admission No. {student} Name: {students_pool[student].get_name()} No. Of. Duties {len(students_pool[student].get_duties())}")

# #########################################################################################################################
# ############################################### Updating Students List ##################################################
# ######################################################################################################################### 

# # print(len(alloted_students_pool))

# # # Loading the Excel File
# # workbook = openpyxl.load_workbook('StudentList.xlsx')

# # # Slecting the First(0th - indexed) Worksheet/Sheet in the Workbook/Excel File.
# # worksheet = workbook.worksheets[0]

# # # Creating a new Workbook and Worksheet
# # new_workbook = openpyxl.Workbook()
# # new_worksheet = new_workbook.active

# # # Copying the Header Row from the Original Worksheet to the new Worksheet
# # for row in worksheet.iter_rows(min_row=1, max_row=1, values_only=True):
# #     new_worksheet.append(row)

# # # Iterating through the rows in the Original Worksheet
# # for row in worksheet.iter_rows(min_row=2, values_only=True):
# #     # Getting the Admission No. from the row
# #     admission_no = row[1]
# #     if admission_no == None:
# #         break
# #     # Checking if the Admission No. is in the dictionary
# #     if admission_no.lower() in alloted_students_pool:
# #         # print(admission_no.lower())
# #         # Creating a new row with the "Previous" column updated to "Yes"
# #         new_row = row[:5] + ('Yes',) + row[6:]
# #         # print("It was here")
# #         # Appending the new row to the new Worksheet
# #         new_worksheet.append(new_row)
# #     else:
# #         # If the Admission No. is not in the dictionary, copy the original row as is
# #         new_worksheet.append(row)

# # # save the changes to the new Excel file
# # new_workbook.save('UpdatedStudentList.xlsx')
# # count = 0
# # for student in alloted_students_pool:
# #     if student not in students_available_copy:
# #         count+=1
# #     else:
# #         print(student)

# # print(count)

sorted_courses = sorted(course_list, key=lambda c: (c.get_date(), c.get_time()))
# # for course in sorted_courses:
# #     print(f"Course Code: {course.get_course_acronym()}")
# #     print(f"Date: {course.get_date().date()}")
# #     print(f"Day: {course.get_day()}")
# #     course_invigilators = course.get_invigilators()
# #     for invigilators in course_invigilators:
# #         print(f"Admission No. {invigilators.get_admission_no()} | Email ID: {invigilators.get_email()} | Name: {invigilators.get_name()}")

# create an empty dataframe with the specified columns
df = pd.DataFrame(columns=['Date', 'Day', 'Time', 'Course Acronym', 'Admission No.', 'Name', 'Email ID'])

# iterate through the sorted courses
for course in sorted_courses:
    date = course.get_date().date()
    day = course.get_day()
    time = course.get_time()
    course_acronym = course.get_course_acronym()
    room_list = course.get_room_no()
    building = course.get_building()
    strength = course.get_strength()

    # print(course_acronym)
    # print(strength)

    # req_list = []
    # if len(room_list) > 1:
    #     for room in room_list:
    #         num_students = min(strength, room_capacity[(room.lower()).strip()])
    #         num_invigilator = get_num_invigilator(num_students)
    #         print(num_invigilator)
    #         req_list.append(num_invigilator)
    #         strength -= num_students

    # iterate through the invigilators for the course
    for invigilator in course.get_invigilators():
        
        
        admission_no = invigilator.get_admission_no()
        name = invigilator.get_name()
        email = invigilator.get_email()
        num_duties = len(invigilator.get_duties())
        


        # add a row to the dataframe for each invigilator
        df = df._append({'Date': date, 'Day': day, 'Time': time, 'Course Acronym': course_acronym,
                        'Admission No.': admission_no, 'Name': name, 'Email ID': email, 'Room_no' : room_list, 'Building' : building, 'No. Of. Duties' : num_duties}, ignore_index=True)

# write the dataframe to an Excel file
df.to_excel('InvigilatorList.xlsx', index=False)
